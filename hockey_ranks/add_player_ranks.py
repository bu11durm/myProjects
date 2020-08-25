import openpyxl
from data.misc.player_rank_dict import player_rank

path = "data\misc\keeper_espn_2021.xlsx"

# read from this spreadsheet
workbook = openpyxl.load_workbook(path)
sheet1 = workbook['ESPN']
save_able = False

try:
    workbook.save(path)
    save_able = True
except:
    print('Is the file open?')

if save_able == True:
    print(str(sheet1.max_row))
    print(str(sheet1.max_column))
    for r in range(3,sheet1.max_row + 1):
        for c in range(2,sheet1.max_column + 1):
            if sheet1.cell(row=r, column=c).value in player_rank:
                new_cell = "(" + player_rank.get(sheet1.cell(row=r, column=c).value) + ") " + sheet1.cell(row=r, column=c).value
                print(new_cell)
                sheet1.cell(row=r, column=c).value = new_cell
            # if sheet1.cell(row=r, column=c).value == None:
            #     print('skip')
            # else:
            #     if sheet1.cell(row=r, column=c).value in player_rank:
            #         print("(" + player_rank.get(sheet1.cell(row=r, column=c).value) + ") " + sheet1.cell(row=r, column=c).value)
            #     else:
            #         print(sheet1.cell(row=r, column=c).value)

    workbook.save(path)
