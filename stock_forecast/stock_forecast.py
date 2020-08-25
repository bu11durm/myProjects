import openpyxl
from selenium import webdriver
import time
import datetime

stock_name_column = 1
stock_rating_column = 2
max_data_days = 50
max_reread_tries = 5

# read from this spreadsheet
path = 'data\stock_forecast.xlsx'
workbook = openpyxl.load_workbook(path)
sheet1 = workbook['DataTable']
save_able = False

try:
    workbook.save('data\stock_forecast.xlsx')
    save_able = True
except:
    print('Is the file open?')

if save_able == True:
    #driver = webdriver.Chrome('venv/chromedriver.exe')
    my_options = webdriver.ChromeOptions()
    my_options.add_argument('--ignore-certificate-errors')
    my_options.add_argument('--ignore-ssl-errors')
    driver = webdriver.Chrome('../venv/chromedriver.exe', options=my_options)
    stock_rating_prefix = 'https://stockinvest.us/technical-analysis/'

    sheet1.delete_cols(max_data_days+1)
    sheet1.insert_cols(stock_rating_column)

    sheet1.cell(row=1, column = stock_rating_column).value = str(datetime.date.today())

    for r in range(2, sheet1.max_row + 1):
        stock_name = sheet1.cell(row=r, column=stock_name_column).value
        driver.get(stock_rating_prefix + stock_name)
        time.sleep(6)
        stock_rating = driver.find_element_by_id('gauge-text').text.split(': ')
        same_read = False
        reread_tries = 1
        while (same_read == False) & (reread_tries < max_reread_tries):
            time.sleep(3)
            stock_rating_2 = driver.find_element_by_id('gauge-text').text.split(': ')
            if stock_rating == stock_rating_2:
                same_read = True
                print('same read is true')
            else:
                stock_rating = stock_rating_2
                print(f'rereading try: {reread_tries}')
                reread_tries += 1
                if reread_tries == max_reread_tries:
                    print('never got same number - using last one')

        try:
            sheet1.cell(row=r, column=stock_rating_column).value = float(stock_rating[1])
        except:
            print(f'error row {r} column {stock_rating_column}')

    workbook.save('data\stock_forecast.xlsx')
    driver.close()


